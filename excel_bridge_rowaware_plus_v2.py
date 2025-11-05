from flask import Flask, request, jsonify
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import datetime, threading

# ==== CONFIG ====
PROJECT_ROOT = Path(__file__).resolve().parent
PRIMARY_WORKBOOK = PROJECT_ROOT / "Excel & Report" / "OU WBB Defensive Project.xlsx"
FALLBACK_WORKBOOK = PROJECT_ROOT / "Excel & Report" / "OU WBB Defensive Project copy.xlsx"
if PRIMARY_WORKBOOK.exists():
    WORKBOOK_PATH = PRIMARY_WORKBOOK
elif FALLBACK_WORKBOOK.exists():
    WORKBOOK_PATH = FALLBACK_WORKBOOK
else:
    WORKBOOK_PATH = PRIMARY_WORKBOOK
SHEET_NAME = "Tagging"
# ================

app = Flask(__name__)
# Avoid sorting JSON keys (prevents None vs str comparison errors)
app.config['JSON_SORT_KEYS'] = False

@app.after_request
def add_cors_headers(resp):
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    resp.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS, GET"
    return resp

_lock = threading.Lock()

def ensure_workbook():
    if WORKBOOK_PATH.exists():
        wb = load_workbook(WORKBOOK_PATH)
    else:
        wb = Workbook()
    if SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(SHEET_NAME)
    return wb

def row_is_empty(ws, r):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=r, column=c).value not in (None, ""):
            return False
    return True

def write_row_to_sheet(ws, row_dict, target_row: int, overwrite: bool = False):
    # Ensure header
    if ws.max_row == 1 and ws.max_column == 1 and (ws["A1"].value is None):
        for j, h in enumerate(row_dict.keys(), start=1):
            ws.cell(row=1, column=j, value=h)
    else:
        headers = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        new = [k for k in row_dict.keys() if k not in headers]
        if new:
            start_col = ws.max_column + 1
            for idx, k in enumerate(new, start=start_col):
                ws.cell(row=1, column=idx, value=k)

    headers = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]

    if overwrite:
        # Write directly to the specified row
        r = max(2, int(target_row))
    else:
        # Find next empty row at or after target
        r = max(2, int(target_row))
        limit = max(ws.max_row + 200, r + 200)
        while r <= limit and not row_is_empty(ws, r):
            r += 1

    for col_index, key in enumerate(headers, start=1):
        ws.cell(row=r, column=col_index, value=row_dict.get(key, ""))

    if r <= 200:
        for j, _ in enumerate(headers, start=1):
            col = get_column_letter(j)
            try:
                max_len = 0
                for rr in range(1, ws.max_row + 1):
                    v = ws.cell(row=rr, column=j).value
                    max_len = max(max_len, len(str(v)) if v else 0)
                ws.column_dimensions[col].width = min(max(10, max_len + 2), 40)
            except Exception:
                pass

    return r

@app.route("/check_row", methods=["GET", "OPTIONS"])
def check_row():
    if request.method == "OPTIONS":
        return jsonify({"ok": True})
    try:
        row_num = int(request.args.get("row", "2"))
        with _lock:
            wb = ensure_workbook()
            ws = wb[SHEET_NAME]
            has_data = not row_is_empty(ws, row_num)
        return jsonify({"ok": True, "has_data": has_data, "row": row_num})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/append", methods=["POST", "OPTIONS"])
def append():
    if request.method == "OPTIONS":
        return jsonify({"ok": True})
    try:
        data = request.get_json(force=True)
        if not isinstance(data, dict):
            return jsonify({"ok": False, "error": "Payload must be a JSON object"}), 400
        data.setdefault("Bridge_Received_At", datetime.datetime.now(datetime.UTC).isoformat())

        target_row = data.pop("Target_Row", None) or data.pop("target_row", None)
        overwrite = data.pop("Overwrite", False) or data.pop("overwrite", False)
        
        try:
            target_row = int(target_row) if target_row is not None else 2
        except Exception:
            target_row = 2

        with _lock:
            wb = ensure_workbook()
            ws = wb[SHEET_NAME]
            used = write_row_to_sheet(ws, data, target_row, overwrite)
            wb.save(WORKBOOK_PATH)

        print(f"✅ Saved to row {used} → {WORKBOOK_PATH.name} / {SHEET_NAME}")
        return jsonify({"ok": True, "saved_to": str(WORKBOOK_PATH), "row": used})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/peek", methods=["GET"])
def peek():
    # Return last N rows (default 3) with safe string headers
    try:
        n = int(request.args.get("rows", "3"))
        with _lock:
            wb = ensure_workbook()
            ws = wb[SHEET_NAME]
            rows = list(ws.values)
            if not rows:
                return jsonify({"ok": True, "rows": [], "count": 0, "sheet": SHEET_NAME, "workbook": str(WORKBOOK_PATH)})
            raw_headers = list(rows[0])
            headers = []
            for idx, h in enumerate(raw_headers, start=1):
                headers.append(str(h) if h is not None else f"col{idx}")
            body = rows[1:]
            last = body[-n:] if n>0 else []
            out = [dict(zip(headers, r)) for r in last]
        return jsonify({"ok": True, "rows": out, "count": len(out), "sheet": SHEET_NAME, "workbook": str(WORKBOOK_PATH)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"ok": True, "workbook": str(WORKBOOK_PATH), "sheet": SHEET_NAME})

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5001, debug=False)
